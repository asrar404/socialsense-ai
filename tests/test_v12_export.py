"""V12 Phase K - Export & Reporting Intelligence tests."""
import csv
import io
import json
import os
from datetime import datetime, timezone
from database import db
from models.media_analysis import MediaAnalysis
from models.threat_assessment import ThreatAssessment
from models.narrative import Narrative
from models.narrative_occurrence import NarrativeOccurrence
from models.coordination_signal import CoordinationSignal
from models.propagation_event import PropagationEvent
from models.entity import Entity
from models.entity_context import EntityContext


def _now():
    return datetime(2024, 1, 1, tzinfo=timezone.utc).replace(tzinfo=None)


def _seed_threat(db, user, analysis_id, score=62.5, level='Elevated'):
    ta = ThreatAssessment(
        analysis_id=analysis_id, user_id=user.id,
        overall_threat_score=score, threat_level=level,
        confidence=72.0, evidence_coverage=0.66, agreement_score=80.0,
        authenticity_score=40.0, coordination_score=30.0,
        narrative_risk_score=62.5, propagation_score=None,
        temporal_score=None, entity_risk_score=None,
        component_scores={'authenticity': 40.0, 'coordination': 30.0,
                          'narrative': 62.5},
        component_weights={'authenticity': 0.25, 'coordination': 0.25,
                           'narrative': 0.20},
        available_components=['authenticity', 'coordination', 'narrative'],
        missing_components=['propagation', 'temporal', 'entity'],
        capability_labels={'authenticity': 'heuristic',
                           'coordination': 'heuristic',
                           'narrative': 'heuristic',
                           'propagation': 'unavailable',
                           'temporal': 'unavailable',
                           'entity': 'unavailable'},
        summary='Heuristic threat assessment.',
        reasons=['Heuristic weighted combination.'],
        indicators=['signal'],
        limitations=['Not a trained ML classifier.'],
        assessment_method='heuristic_weighted',
    )
    db.session.add(ta)
    db.session.commit()
    return ta


def _seed_media_analysis(db, analysis_id):
    ma = MediaAnalysis(
        analysis_id=analysis_id, overall_ai_probability=55.0,
        overall_authenticity_score=45.0, confidence=70.0,
        deepfake_score=50.0, synthetic_voice_score=55.0,
        thumbnail_ai_score=50.0, frame_manipulation_score=45.0,
        metadata_score=40.0,
        summary='Heuristic media authenticity assessment.',
        reasons='["Sample reason one", "Sample reason two"]',
    )
    db.session.add(ma)
    db.session.commit()
    return ma


def _seed_narrative(db, user, analysis_id, name='Election fraud claims'):
    narrative = Narrative(
        user_id=user.id, name=name, normalized_name=name.lower(),
        category='political', risk_score=62.5, confidence=0.7,
        growth_score=45.0, occurrence_count=2, platform_count=2,
        keywords=['fraud', 'election'], entity_names=['PERSON_A'],
        evidence={'sources': ['youtube', 'reddit'],
                  'samples': ['sample one', 'sample two']},
        detection_method='heuristic', first_seen_at=_now(),
        last_seen_at=_now(),
    )
    db.session.add(narrative)
    db.session.flush()
    db.session.add(NarrativeOccurrence(
        narrative_id=narrative.id, analysis_id=analysis_id,
        user_id=user.id, platform='youtube', source='comment',
        relevance_score=0.8, risk_score=62.5, match_count=3,
        evidence={'samples': ['sample one', 'sample two']},
        occurred_at=_now(),
    ))
    db.session.commit()
    return narrative


def _seed_coordination(db, user, analysis_id):
    signal = CoordinationSignal(
        analysis_id=analysis_id, user_id=user.id,
        signal_type='repeated_content', score=58.0, confidence=0.65,
        level='elevated', cluster_size=4, comparisons_performed=120,
        comparisons_truncated=True, window_seconds=3600,
        summary='Repeated phrasing across comments.',
        reasons=['Repeated content'], indicators=['repetition'],
        evidence={'sample_text': 'repeat'}, related_entities=['PERSON_A'],
        detection_method='heuristic', detected_at=_now(),
    )
    db.session.add(signal)
    db.session.commit()
    return signal


def _seed_propagation(db, user, source_id, target_id):
    event = PropagationEvent(
        user_id=user.id, narrative_id=None,
        source_analysis_id=source_id, target_analysis_id=target_id,
        source_platform='youtube', target_platform='reddit',
        source_ref='vid1', target_ref='post1',
        relationship_type='potentially_propagated',
        direction='source_to_target', propagation_score=55.0,
        confidence=0.6, similarity_score=0.7, lag_seconds=7200,
        shared_entities=['PERSON_A'], reasons=['Shared entities and timing'],
        evidence={'samples': ['sample']}, detection_method='heuristic',
        occurred_at=_now(),
    )
    db.session.add(event)
    db.session.commit()
    return event


def _seed_entity(db, analysis_id):
    from models.comment_result import CommentResult
    comment = CommentResult.query.filter_by(analysis_id=analysis_id).first()
    ent = Entity(analysis_id=analysis_id, name='PERSON_A',
                 normalized_name='person_a', entity_type='PERSON',
                 source='comment', frequency=5, importance_score=60.0)
    db.session.add(ent)
    db.session.flush()
    ec = EntityContext(entity_id=ent.id, comment_result_id=comment.id,
                       entity_sentiment='positive',
                       entity_sentiment_score=65.0, entity_risk_score=30.0)
    db.session.add(ec)
    db.session.commit()
    return ent


def _seed_full_v12(db, user, analysis):
    _seed_threat(db, user, analysis.id)
    _seed_media_analysis(db, analysis.id)
    _seed_narrative(db, user, analysis.id)
    _seed_coordination(db, user, analysis.id)
    _seed_propagation(db, user, analysis.id, analysis.id + 1)
    _seed_entity(db, analysis.id)
# --------------------------------------------------------------------------
# CSV
# --------------------------------------------------------------------------
class TestCsvExport:
    def test_route_requires_login(self, client, analysis):
        response = client.get(f'/export/csv/{analysis.id}')
        assert response.status_code == 302

    def test_route_success(self, logged_in_client, analysis):
        response = logged_in_client.get(f'/export/csv/{analysis.id}')
        assert response.status_code == 200
        assert 'text/csv' in response.mimetype

    def test_v12_threat_in_csv(self, app, db, user, analysis, logged_in_client):
        _seed_threat(db, user, analysis.id)
        response = logged_in_client.get(f'/export/csv/{analysis.id}')
        content = response.data.decode('utf-8')
        assert 'Overall Threat Score' in content
        assert '62.5' in content
        assert 'Elevated' in content

    def test_v12_narratives_in_csv(self, app, db, user, analysis,
                                   logged_in_client):
        _seed_narrative(db, user, analysis.id)
        response = logged_in_client.get(f'/export/csv/{analysis.id}')
        content = response.data.decode('utf-8')
        assert 'Election fraud claims' in content
        assert 'political' in content

    def test_v12_coordination_in_csv(self, app, db, user, analysis,
                                     logged_in_client):
        _seed_coordination(db, user, analysis.id)
        response = logged_in_client.get(f'/export/csv/{analysis.id}')
        content = response.data.decode('utf-8')
        assert 'repeated_content' in content
        assert '120' in content

    def test_v12_propagation_in_csv(self, app, db, user, analysis,
                                    logged_in_client):
        _seed_propagation(db, user, analysis.id, analysis.id + 1)
        response = logged_in_client.get(f'/export/csv/{analysis.id}')
        content = response.data.decode('utf-8')
        assert 'youtube' in content
        assert 'reddit' in content

    def test_evidence_in_csv(self, app, db, user, analysis, logged_in_client):
        _seed_media_analysis(db, analysis.id)
        _seed_narrative(db, user, analysis.id)
        response = logged_in_client.get(f'/export/csv/{analysis.id}')
        content = response.data.decode('utf-8')
        assert 'Sample reason one' in content
        assert 'sample one' in content

    def test_existing_content_preserved(self, logged_in_client, analysis):
        response = logged_in_client.get(f'/export/csv/{analysis.id}')
        content = response.data.decode('utf-8')
        assert 'Video ID' in content
        assert 'dQw4w9WgXcQ' in content
        assert 'Comment' in content
        assert 'Risk Score' in content

    def test_not_found(self, logged_in_client):
        response = logged_in_client.get('/export/csv/999')
        assert response.status_code == 302

    def test_unauthorized(self, app, analysis):
        from services.export_service import ExportService
        result = ExportService().generate_csv(analysis.id, 999)
        assert result is None


# --------------------------------------------------------------------------
# JSON
# --------------------------------------------------------------------------
class TestJsonExport:
    def test_route_requires_login(self, client, analysis):
        response = client.get(f'/export/json/{analysis.id}')
        assert response.status_code == 302

    def test_route_success(self, logged_in_client, analysis):
        response = logged_in_client.get(f'/export/json/{analysis.id}')
        assert response.status_code == 200
        assert 'application/json' in response.mimetype

    def test_v12_key_present(self, app, db, user, analysis, logged_in_client):
        _seed_full_v12(db, user, analysis)
        response = logged_in_client.get(f'/export/json/{analysis.id}')
        data = json.loads(response.data)
        assert 'v12' in data
        assert 'threat' in data['v12']
        assert 'narratives' in data['v12']
        assert 'coordination' in data['v12']
        assert 'propagation' in data['v12']
        assert 'temporal' in data['v12']
        assert 'heuristic_disclaimer' in data['v12']
        assert 'threat_disclaimer' in data['v12']

    def test_entities_key_present(self, app, db, user, analysis,
                                  logged_in_client):
        _seed_entity(db, analysis.id)
        response = logged_in_client.get(f'/export/json/{analysis.id}')
        data = json.loads(response.data)
        assert 'entities' in data
        assert len(data['entities']) > 0

    def test_full_v12_json(self, app, db, user, analysis, logged_in_client):
        _seed_full_v12(db, user, analysis)
        response = logged_in_client.get(f'/export/json/{analysis.id}')
        data = json.loads(response.data)
        assert data['v12']['threat']['overall_threat_score'] == 62.5
        assert data['v12']['threat']['threat_level'] == 'Elevated'
        assert len(data['v12']['narratives']) > 0
        assert len(data['v12']['coordination']) > 0
        assert len(data['v12']['propagation']) > 0

    def test_empty_threat_never_zero(self, app, db, user, analysis,
                                     logged_in_client):
        response = logged_in_client.get(f'/export/json/{analysis.id}')
        data = json.loads(response.data)
        v12 = data.get('v12', {})
        if v12.get('threat'):
            score = v12['threat'].get('overall_threat_score')
            assert score is None, 'empty threat should not fake a score'

    def test_existing_json_content_preserved(self, logged_in_client, analysis):
        response = logged_in_client.get(f'/export/json/{analysis.id}')
        data = json.loads(response.data)
        assert 'analysis_id' in data
        assert 'analysis_type' in data
        assert 'comments' in data
        assert 'comment_count' in data

    def test_not_found(self, logged_in_client):
        response = logged_in_client.get('/export/json/999')
        assert response.status_code == 302

    def test_unauthorized(self, app, analysis):
        from services.export_service import ExportService
        result = ExportService().generate_json(analysis.id, 999)
        assert result is None


# --------------------------------------------------------------------------
# XLSX
# --------------------------------------------------------------------------
class TestXlsxExport:
    def test_route_requires_login(self, client, analysis):
        response = client.get(f'/export/xlsx/{analysis.id}')
        assert response.status_code == 302

    def test_route_success(self, logged_in_client, analysis):
        response = logged_in_client.get(f'/export/xlsx/{analysis.id}')
        assert response.status_code == 200
        assert 'spreadsheet' in response.mimetype

    def test_sheet_structure(self, app, db, user, analysis):
        _seed_full_v12(db, user, analysis)
        from services.export_service import ExportService
        from openpyxl import load_workbook
        result = ExportService().generate_xlsx(analysis.id, user.id)
        assert result is not None
        wb = load_workbook(result['filepath'])
        actual = set(wb.sheetnames)
        for name in ('Summary', 'Threat Assessment', 'Authenticity Intelligence',
                     'Narratives & Growth', 'Coordination Signals',
                     'Propagation Relationships', 'Temporal Intelligence',
                     'Entity Intelligence', 'Evidence'):
            assert name in actual, f'Missing sheet: {name}'
        wb.close()
        os.remove(result['filepath'])

    def test_xlsx_content(self, app, db, user, analysis):
        _seed_threat(db, user, analysis.id)
        from services.export_service import ExportService
        from openpyxl import load_workbook
        result = ExportService().generate_xlsx(analysis.id, user.id)
        assert result is not None
        wb = load_workbook(result['filepath'])
        ws = wb['Threat Assessment']
        found = any(cell.value and '62.5' in str(cell.value)
                    for row in ws.iter_rows() for cell in row)
        assert found
        wb.close()
        os.remove(result['filepath'])

    def test_empty_xlsx_has_disclaimers(self, app, db, user, analysis):
        from services.export_service import ExportService
        from openpyxl import load_workbook
        result = ExportService().generate_xlsx(analysis.id, user.id)
        assert result is not None
        wb = load_workbook(result['filepath'])
        content = ' '.join(str(cell.value or '')
                           for ws in wb.worksheets
                           for row in ws.iter_rows() for cell in row)
        assert 'heuristic' in content.lower() or 'non-causal' in content.lower()
        wb.close()
        os.remove(result['filepath'])

    def test_not_found(self, logged_in_client):
        response = logged_in_client.get('/export/xlsx/999')
        assert response.status_code == 302

    def test_unauthorized(self, app, analysis):
        from services.export_service import ExportService
        result = ExportService().generate_xlsx(analysis.id, 999)
        assert result is None
# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------
class TestPdfExport:
    def test_route_requires_login(self, client, analysis):
        response = client.get(f'/export/pdf/{analysis.id}')
        assert response.status_code == 302

    def test_route_success(self, logged_in_client, analysis):
        response = logged_in_client.get(f'/export/pdf/{analysis.id}')
        assert response.status_code == 200
        assert 'application/pdf' in response.mimetype

    def test_pdf_generate_full(self, app, db, user, analysis):
        _seed_full_v12(db, user, analysis)
        from services.export_service import ExportService
        result = ExportService().generate_pdf(analysis.id, user.id)
        assert result is not None
        assert result['filename'].endswith('.pdf')
        assert os.path.getsize(result['filepath']) > 0
        os.remove(result['filepath'])

    def test_pdf_generate_empty(self, app, db, user, analysis):
        from services.export_service import ExportService
        result = ExportService().generate_pdf(analysis.id, user.id)
        assert result is not None
        assert os.path.getsize(result['filepath']) > 0
        os.remove(result['filepath'])

    def test_not_found(self, logged_in_client):
        response = logged_in_client.get('/export/pdf/999')
        assert response.status_code == 302

    def test_unauthorized(self, app, analysis):
        from services.export_service import ExportService
        result = ExportService().generate_pdf(analysis.id, 999)
        assert result is None


# --------------------------------------------------------------------------
# DOCX
# --------------------------------------------------------------------------
class TestDocxExport:
    def test_route_requires_login(self, client, analysis):
        response = client.get(f'/export/docx/{analysis.id}')
        assert response.status_code == 302

    def test_route_success(self, logged_in_client, analysis):
        response = logged_in_client.get(f'/export/docx/{analysis.id}')
        assert response.status_code == 200
        assert 'document' in response.mimetype

    def test_docx_generate_full(self, app, db, user, analysis):
        _seed_full_v12(db, user, analysis)
        from services.export_service import ExportService
        from docx import Document
        result = ExportService().generate_docx(analysis.id, user.id)
        assert result is not None
        assert result['filename'].endswith('.docx')
        doc = Document(result['filepath'])
        text = '\n'.join(p.text for p in doc.paragraphs)
        assert 'Threat Assessment' in text
        assert 'Narratives & Growth' in text
        assert 'heuristic' in text.lower()
        os.remove(result['filepath'])

    def test_docx_generate_empty(self, app, db, user, analysis):
        from services.export_service import ExportService
        result = ExportService().generate_docx(analysis.id, user.id)
        assert result is not None
        os.remove(result['filepath'])

    def test_not_found(self, logged_in_client):
        response = logged_in_client.get('/export/docx/999')
        assert response.status_code == 302

    def test_unauthorized(self, app, analysis):
        from services.export_service import ExportService
        result = ExportService().generate_docx(analysis.id, 999)
        assert result is None


# --------------------------------------------------------------------------
# Cross-format & regression
# --------------------------------------------------------------------------
class TestCrossFormat:
    def test_all_formats_route_ok(self, app, db, user, analysis,
                                  logged_in_client):
        _seed_full_v12(db, user, analysis)
        for fmt in ('csv', 'json', 'xlsx', 'pdf', 'docx'):
            response = logged_in_client.get(f'/export/{fmt}/{analysis.id}')
            assert response.status_code == 200, fmt

    def test_reddit_csv_backwards_compatible(self, logged_in_client,
                                             reddit_analysis):
        response = logged_in_client.get(f'/export/csv/{reddit_analysis.id}')
        assert response.status_code == 200
        content = response.data.decode('utf-8')
        assert 'Post ID' in content
        assert 'abc123' in content
        assert 'Test Reddit Post' in content

    def test_reddit_xlsx(self, logged_in_client, reddit_analysis):
        from services.export_service import ExportService
        from openpyxl import load_workbook
        result = ExportService().generate_xlsx(
            reddit_analysis.id, reddit_analysis.user_id)
        assert result is not None
        wb = load_workbook(result['filepath'])
        assert 'Summary' in wb.sheetnames
        wb.close()
        os.remove(result['filepath'])

    def test_json_disclaimers_present(self, app, db, user, analysis,
                                      logged_in_client):
        _seed_full_v12(db, user, analysis)
        response = logged_in_client.get(f'/export/json/{analysis.id}')
        data = json.loads(response.data)
        disclaimers = [data['v12']['heuristic_disclaimer'],
                       data['v12']['threat_disclaimer']]
        assert all(disclaimers)
        joined = json.dumps(disclaimers).lower()
        assert 'heuristic' in joined
        assert 'non-causal' in joined

    def test_export_failure_does_not_crash_session(self, app, analysis,
                                                   logged_in_client):
        response = logged_in_client.get('/export/json/999')
        assert response.status_code == 302
        assert logged_in_client.get('/analysis/history').status_code == 200
